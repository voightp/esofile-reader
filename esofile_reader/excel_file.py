import os
from datetime import datetime
from pathlib import Path
from typing import Union, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from esofile_reader.base_file import BaseFile
from esofile_reader.constants import *
from esofile_reader.data.df_data import DFData
from esofile_reader.exceptions import InsuficientHeaderInfo
from esofile_reader.processor.monitor import DefaultMonitor
from esofile_reader.search_tree import Tree


def is_data_row(sr: pd.Series):
    """ Check if the given series is header or data row. """

    def check(val):
        if pd.isna(val):
            return val
        return "num" if isinstance(val, (int, float, np.float, np.int)) else "not_num"

    num_count = sr.apply(check).value_counts(dropna=False)

    # it's difficult to generalize if the row is numeric for mixed types
    # let's try just verifying if there's more numeric than non numeric
    # columns, excluding nan
    numeric_count = num_count["num"] if "num" in num_count else 0
    non_numeric_count = num_count["not_num"] if "not_num" in num_count else 0
    nat_count = num_count[pd.NaT] if pd.NaT in num_count else 0

    return numeric_count >= non_numeric_count


class ExcelFile(BaseFile):
    """ Create results file based on excel data. """

    HEADER_LIMIT = 10

    def __init__(
            self,
            file_path: Union[str, Path],
            sheet_names: List[str] = None,
            force_index: bool = False,
            monitor: DefaultMonitor = None
    ):
        super().__init__()
        self.file_path = file_path
        self.populate_content(monitor, sheet_names=sheet_names, force_index=force_index)

    @staticmethod
    def parse_header(
            df: pd.DataFrame, force_index: bool = False
    ) -> Tuple[pd.MultiIndex, int, bool]:
        """ Extract header related information from excel worksheet. """
        index_names = [TIMESTAMP_COLUMN, RANGE, INDEX]
        column_levels = [
            ID_LEVEL,
            INTERVAL_LEVEL,
            KEY_LEVEL,
            TYPE_LEVEL,
            UNITS_LEVEL,
        ]
        first_column = df.iloc[:, 0].tolist()

        # try to find out if DataFrame includes index column
        if force_index:
            index_column = True
        else:
            # check if DataFrame include index column
            index_column = False
            for cell in first_column:
                conditions = [
                    cell in index_names,
                    isinstance(cell, (datetime, np.datetime64)),
                ]
                if any(conditions):
                    index_column = True
                    break

        # check if DataFrame has a 'template' structure
        if index_column and KEY_LEVEL in first_column and UNITS_LEVEL in first_column:
            is_template = True
        else:
            is_template = False

        levels = {}
        i = 0
        for _, sr in df.iterrows():
            if index_column:
                ix = sr.iloc[0]
                row = sr.iloc[1:]
                if row.dropna(how="all").empty:
                    # ignore rows without any value
                    pass
                elif is_template:
                    if ix in index_names:
                        # hit either RANGE or TIMESTAMP or INDEX keyword
                        i += 1
                        break
                    elif isinstance(ix, (datetime, int)):
                        # hit integer range or datetime timestamp
                        break
                    elif ix in column_levels:
                        if ix == ID_LEVEL:
                            # ID will be re-generated to avoid necessary validation
                            pass
                        else:
                            levels[ix] = row
                    else:
                        print(
                            f"Unexpected column identifier: {ix}"
                            f"Only {', '.join(COLUMN_LEVELS)} are allowed."
                        )
                else:
                    if is_data_row(row):
                        # hit actual data rows
                        break
                    else:
                        levels[i] = row
            else:
                if is_data_row(sr):
                    break
                levels[i] = sr.values
            i += 1
        else:
            raise InsuficientHeaderInfo(
                "Failed to automatically retrieve header information from DataFrame!"
            )

        # validate gathered data
        ordered_levels = {}
        if is_template:
            if KEY_LEVEL not in levels or UNITS_LEVEL not in levels:
                raise InsuficientHeaderInfo(
                    f"Cannot process header!"
                    f" '{KEY_LEVEL}' and '{UNITS_LEVEL}' levels must be included."
                )
            # reorder levels using standard order
            ordered_levels = {lev: levels[lev] for lev in column_levels if lev in levels}
        else:
            n = len(levels)
            msg = (
                " expected levels are either:\n\t1 - key\n\tunits"
                "\n..for two level header or:\n\t1 - key\n\ttype\nunits"
                " for three level header."
            )
            if n < 2:
                raise InsuficientHeaderInfo(
                    f"Not enough information to create header. " f"There's only {n} but {msg}"
                )
            elif n > 3:
                raise InsuficientHeaderInfo(f"Too many header levels - {n}\n{msg}")
            keys = [KEY_LEVEL, TYPE_LEVEL, UNITS_LEVEL]
            if n == 2:
                keys.remove(TYPE_LEVEL)
            for key, row in zip(keys, list(levels.values())):
                ordered_levels[key] = row

        # transpose DataFrame to get items in columns
        header_df = pd.DataFrame(ordered_levels).T

        # replace 'Nan' values with empty strings
        header_df.fillna(value="", inplace=True)

        header_mi = pd.MultiIndex.from_frame(header_df.T, names=list(header_df.index))

        return header_mi, i, index_column

    @staticmethod
    def build_df_table(
            raw_df: pd.DataFrame, name: str, start_id: int = 1,
    ) -> Tuple[pd.DataFrame, int]:
        """ Finalize DataFrame data to match required DFData structure. """
        # include table name row if it's not already present
        if INTERVAL_LEVEL not in raw_df.columns.names:
            raw_df = pd.concat([raw_df], keys=[name], names=[INTERVAL_LEVEL], axis=1)

        key_level = raw_df.columns.get_level_values(KEY_LEVEL)
        special_rows = key_level.isin([DAY_COLUMN, N_DAYS_COLUMN])
        # all special columns have id 'special'
        special_df = raw_df.loc[:, special_rows]
        special_df = pd.concat([special_df], keys=[SPECIAL], names=[ID_LEVEL], axis=1)

        # create unique ids for each variable
        numeric_df = raw_df.loc[:, ~special_rows]
        end_id = start_id + len(numeric_df.columns)
        header = numeric_df.columns.to_frame()
        header.insert(0, ID_LEVEL, list(range(start_id, end_id)))
        numeric_df.columns = pd.MultiIndex.from_frame(header)

        # merge DataFrames back together
        df = pd.concat([special_df, numeric_df], axis=1)

        # store names as convert_dtypes resets column names!
        names = df.columns.names

        # all columns use 'object' dtype,
        df = df.convert_dtypes()

        # update column and index names
        df.columns.rename(names, inplace=True)
        if isinstance(df.index, pd.DatetimeIndex):
            df.index = df.index.round(freq="S")
            df.index.rename(TIMESTAMP_COLUMN, inplace=True)
        elif isinstance(df.index, pd.RangeIndex):
            df.index.rename(RANGE, inplace=True)
        else:
            df.index.rename(INDEX, inplace=True)

        return df, end_id

    def process_excel(
            self,
            file_path: Union[str, Path],
            sheet_names: List[str] = None,
            force_index: bool = False,
    ):
        """ Create results file data based on given excel workbook."""
        wb = load_workbook(filename=file_path, read_only=True)
        if not sheet_names:
            sheet_names = wb.sheetnames

        start_id = 1
        df_data = DFData()
        for name in sheet_names:
            ws = wb[name]
            df = pd.DataFrame(ws.values)

            # ignore empty rows and columns
            df.dropna(axis=0, how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)

            header_mi, skiprows, index_column = self.parse_header(
                df.iloc[: self.HEADER_LIMIT, :], force_index=force_index
            )

            df = df.iloc[skiprows:, :]
            if index_column:
                df.set_index(keys=df.columns[0], inplace=True)
            else:
                df.reset_index(inplace=True, drop=True)

            df.columns = header_mi

            # create table for each interval name
            if INTERVAL_LEVEL in df:
                interval_level = df.columns.get_level_values(INTERVAL_LEVEL)
                for key in interval_level.unique():
                    dfi, end_id = self.build_df_table(
                        df.loc[:, interval_level == key],
                        name=key,
                        start_id=start_id,
                    )
                    start_id = end_id
                    df_data.populate_table(INTERVAL_LEVEL, dfi)
            else:
                df, _ = self.build_df_table(df, name=name)
                df_data.populate_table(name, df)

        tree = Tree()
        tree.populate_tree(df_data.get_all_variables_dct())

        return df_data, tree

    def populate_content(
            self, monitor: DefaultMonitor = None, sheet_names: str = None,
            force_index: bool = False
    ) -> None:
        self.file_name = Path(self.file_path).stem
        self.file_created = datetime.utcfromtimestamp(os.path.getctime(self.file_path))
        self.data, self.search_tree = self.process_excel(
            self.file_path, sheet_names=sheet_names, force_index=force_index
        )
