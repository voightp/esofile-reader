import csv
import io
import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Dict, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook

from esofile_reader.df.df_tables import DFTables
from esofile_reader.df.level_names import (
    RANGE,
    INDEX,
    N_DAYS_COLUMN,
    DAY_COLUMN,
    TIMESTAMP_COLUMN,
    ID_LEVEL,
    TABLE_LEVEL,
    KEY_LEVEL,
    TYPE_LEVEL,
    UNITS_LEVEL,
    COLUMN_LEVELS,
    SPECIAL,
)
from esofile_reader.exceptions import InsuficientHeaderInfo, NoResults
from esofile_reader.id_generator import get_unique_name
from esofile_reader.processing.progress_logger import BaseLogger


def is_data_row(sr: pd.Series):
    """ Check if the given series is header or data row. """

    def check(val):
        if pd.isna(val):
            return val
        return "num" if isinstance(val, (int, float, np.float, np.int)) else "not_num"

    num_count = sr.apply(check).value_counts(dropna=False)

    # it's difficult to generalize if the row is numeric for mixed types
    # let's try just verifying if there's more numeric than non numeric
    # columns, excluding nan
    numeric_count = num_count["num"] if "num" in num_count else 0
    non_numeric_count = num_count["not_num"] if "not_num" in num_count else 0
    # nat_count = num_count[pd.NaT] if pd.NaT in num_count else 0

    return numeric_count >= non_numeric_count


def validate_header(is_template, levels, column_levels):
    """ Validate collected header information and return ordered header. """
    ordered_levels = {}
    if is_template:
        # reorder levels using standard order
        ordered_levels = {lev: levels[lev] for lev in column_levels if lev in levels}
    else:
        n = len(levels)
        msg = (
            "Expected levels are either:\n\tkey\n\tunits"
            "\nfor two level header or:\n\tkey\n\ttype\n\tunits"
            "\nfor three level header."
        )
        if n < 2:
            raise InsuficientHeaderInfo(
                f"Not enough information to create header. " f"There's only {n} level.\n{msg}"
            )
        elif n > 3:
            raise InsuficientHeaderInfo(f"There's too many header levels: {n}.\n{msg}")
        keys = [KEY_LEVEL, TYPE_LEVEL, UNITS_LEVEL]
        if n == 2:
            keys.remove(TYPE_LEVEL)
        for key, row in zip(keys, list(levels.values())):
            ordered_levels[key] = row
    return ordered_levels


def parse_header(
    df: pd.DataFrame, force_index: bool = False
) -> Tuple[pd.MultiIndex, int, bool]:
    """ Extract header related information from excel worksheet. """
    index_names = [TIMESTAMP_COLUMN, RANGE, INDEX]
    column_levels = [
        ID_LEVEL,
        TABLE_LEVEL,
        KEY_LEVEL,
        TYPE_LEVEL,
        UNITS_LEVEL,
    ]
    first_column = df.iloc[:, 0].tolist()

    # try to find out if DataFrame includes index column
    if force_index:
        index_column = True
    else:
        # check if DataFrame include index column
        index_column = False
        for cell in first_column:
            conditions = [
                cell in index_names,
                isinstance(cell, (datetime, np.datetime64)),
            ]
            if any(conditions):
                index_column = True
                break

    # check if DataFrame has a 'template' structure
    if index_column and KEY_LEVEL in first_column and UNITS_LEVEL in first_column:
        is_template = True
    else:
        is_template = False

    levels = {}
    i = 0
    for _, sr in df.iterrows():
        if index_column:
            ix = sr.iloc[0]
            row = sr.iloc[1:]
            if row.dropna(how="all").empty and ix not in index_names:
                # ignore rows without any value
                pass
            elif is_template:
                if ix in index_names:
                    # hit either RANGE or TIMESTAMP or INDEX keyword
                    i += 1
                    break
                elif isinstance(ix, (datetime, int)):
                    # hit integer range or datetime timestamp
                    break
                elif ix in column_levels:
                    if ix == ID_LEVEL:
                        # ID will be re-generated to avoid necessary validation
                        pass
                    else:
                        levels[ix] = row
                else:
                    logging.info(
                        f"Unexpected column identifier: {ix}"
                        f"Only {', '.join(COLUMN_LEVELS)} are allowed."
                    )
            else:
                if is_data_row(row):
                    # hit actual data rows
                    break
                else:
                    levels[i] = row
        else:
            if is_data_row(sr):
                break
            levels[i] = sr.values
        i += 1
    else:
        raise InsuficientHeaderInfo(
            "Failed to automatically retrieve header information from DataFrame!"
        )

    # validate gathered information, raises 'InsufficientHeaderInfo'
    ordered_levels = validate_header(is_template, levels, column_levels)

    # transpose DataFrame to get items in columns
    header_df = pd.DataFrame(ordered_levels).T

    # replace 'Nan' values with empty strings
    header_df.fillna(value="", inplace=True)

    header_mi = pd.MultiIndex.from_frame(header_df.T, names=list(header_df.index))

    return header_mi, i, index_column


def build_df_table(
    raw_df: pd.DataFrame, table_name: str, start_id: int = 1,
) -> Tuple[pd.DataFrame, int]:
    """ Finalize DataFrame data to match required DFTables structure. """

    def is_range(array):
        if all(map(lambda x: isinstance(x, int), array)) and len(array) > 1:
            return len(set(np.diff(array))) == 1

    # drop duplicate column items
    duplicated = raw_df.columns.duplicated()
    if any(duplicated):
        raw_df = raw_df.loc[:, ~duplicated]

    # original table name may differ as it may not be unique
    if TABLE_LEVEL in raw_df.columns.names:
        raw_df.columns = raw_df.columns.droplevel(TABLE_LEVEL)
    raw_df = pd.concat([raw_df], keys=[table_name], names=[TABLE_LEVEL], axis=1)

    key_level = raw_df.columns.get_level_values(KEY_LEVEL)
    special_rows = key_level.isin([DAY_COLUMN, N_DAYS_COLUMN])
    # all special columns have id 'special'
    special_df = raw_df.loc[:, special_rows]
    special_df = pd.concat([special_df], keys=[SPECIAL], names=[ID_LEVEL], axis=1)

    # create unique ids for each variable
    numeric_df = raw_df.loc[:, ~special_rows]
    end_id = start_id + len(numeric_df.columns)
    header = numeric_df.columns.to_frame()
    header.insert(0, ID_LEVEL, list(range(start_id, end_id)))
    numeric_df.columns = pd.MultiIndex.from_frame(header)

    # merge DataFrames back together
    df = pd.concat([special_df, numeric_df], axis=1)

    # store names as convert_dtypes resets column names!
    names = df.columns.names

    # all columns use 'object' dtype,
    df = df.convert_dtypes()

    # update column and index names
    df.columns.rename(names, inplace=True)
    if isinstance(df.index, pd.DatetimeIndex):
        df.index = df.index.round(freq="S")
        df.index.rename(TIMESTAMP_COLUMN, inplace=True)
    elif isinstance(df.index, pd.RangeIndex):
        df.index.rename(RANGE, inplace=True)
    elif is_range(df.index):
        # is_range requires range like array with length > 1
        df.index = pd.RangeIndex(
            start=df.index[0], stop=df.index[-1] + 1, step=df.index[1] - df.index[0], name=RANGE
        )
    else:
        df.index.rename(INDEX, inplace=True)

    return df, end_id


def process_sheet(
    df: pd.DataFrame,
    name: str,
    names: List[str],
    logger,
    start_id: int,
    header_limit: int,
    force_index: bool,
) -> Tuple[Dict[str, pd.DataFrame], int]:
    # ignore empty rows and columns
    df.dropna(axis=0, how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    frames = {}
    end_id = start_id
    if not df.empty:
        # process header data
        logger.log_section(f"sheet {name} - processing data dictionary")
        header_mi, skiprows, index_column = parse_header(
            df.iloc[:header_limit, :], force_index=force_index
        )

        # process numeric data
        logger.log_section(f"sheet {name} - processing data")
        df = df.iloc[skiprows:, :]
        if index_column:
            df.set_index(keys=df.columns[0], inplace=True)
        else:
            df.reset_index(inplace=True, drop=True)

        df.columns = header_mi

        logger.log_section(f"sheet {name} - processing tables")
        if TABLE_LEVEL in df.columns.names:
            table_level = df.columns.get_level_values(TABLE_LEVEL)
            for key in table_level.unique():
                table_name = get_unique_name(key, names, start_i=1)
                dfi, end_id = build_df_table(
                    df.loc[:, table_level == key], table_name=table_name, start_id=start_id,
                )
                start_id = end_id
                # ignore empty frames
                if not dfi.loc[:, dfi.columns.get_level_values(ID_LEVEL) != SPECIAL].empty:
                    frames[table_name] = dfi
        else:
            table_name = get_unique_name(name, names, start_i=1)
            df, end_id = build_df_table(df, table_name=table_name, start_id=start_id)
            if not df.loc[:, df.columns.get_level_values(ID_LEVEL) != SPECIAL].empty:
                frames[table_name] = df
    return frames, end_id


def process_workbook(
    wb: Workbook,
    logger: BaseLogger,
    sheet_names: List[str] = None,
    force_index: bool = False,
    header_limit: int = 10,
) -> DFTables:
    if not sheet_names:
        sheet_names = wb.sheetnames

    # each table represents a single step + add one for tree generation
    n_steps = len(sheet_names) + 1
    logger.log_section("processing sheets")
    logger.set_maximum_progress(n_steps)

    start_id = 1
    df_tables = DFTables()
    for name in sheet_names:
        ws = wb[name]
        frames, end_id = process_sheet(
            df=pd.DataFrame(ws.values),
            name=name,
            names=list(df_tables.keys()),
            logger=logger,
            start_id=start_id,
            header_limit=header_limit,
            force_index=force_index,
        )
        start_id = end_id
        df_tables.extend(frames)
        logger.increment_progress()

    return df_tables


def process_excel(
    file_path: Path,
    logger: BaseLogger,
    sheet_names: List[str] = None,
    force_index: bool = False,
    header_limit: int = 10,
) -> DFTables:
    """ Create results file data based on given excel workbook."""
    with open(file_path, "rb") as f:
        in_memory_file = io.BytesIO(f.read())
    wb = load_workbook(filename=in_memory_file, read_only=True)
    return process_workbook(wb, logger, sheet_names, force_index, header_limit)


def process_csv_table(
    df: pd.DataFrame,
    name: str,
    logger: BaseLogger,
    force_index: bool = False,
    header_limit: int = 10,
) -> DFTables:
    """ Process csv sheet. """
    logger.log_section("processing csv")
    df_tables = DFTables()
    frames, _ = process_sheet(
        df=df,
        name=name,
        names=list(df_tables.keys()),
        logger=logger,
        start_id=1,
        header_limit=header_limit,
        force_index=force_index,
    )
    df_tables.extend(frames)
    return df_tables


def process_csv(
    file_path: Path,
    logger: BaseLogger,
    force_index: bool = False,
    header_limit: int = 10,
    separator: Optional[str] = None,
) -> DFTables:
    """ Create results file data based on given csv file."""
    try:
        csv_df = pd.read_csv(file_path, sep=separator, engine="python")
        name = file_path.stem
    except csv.Error:
        raise NoResults(f"Cannot read {file_path}. {traceback.format_exc()}")
    return process_csv_table(csv_df, name, logger, force_index, header_limit)
